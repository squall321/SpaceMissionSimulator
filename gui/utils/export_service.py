"""
ExportService — 분석 결과를 Excel / PDF 로 내보내기
"""
from __future__ import annotations
from pathlib import Path
from datetime import datetime


# ── 공통 행 정의 (ComparisonDialog와 동일) ────────────────────────────────────
EXPORT_ROWS = [
    ("Orbit Type",        lambda r: str(r['orbit'].params.orbit_type)),
    ("Altitude (km)",     lambda r: f"{r['orbit'].params.altitude_km:.1f}"),
    ("Inclination (°)",   lambda r: f"{r['orbit'].params.inclination_deg:.1f}"),
    ("RAAN (°)",          lambda r: f"{r['orbit'].params.raan_deg:.1f}"),
    ("Period (min)",      lambda r: f"{r['orbit'].period_min:.1f}"),
    ("Sunlight (%)",      lambda r: f"{r['orbit'].sunlight_fraction * 100:.1f}"),
    ("Max Eclipse (min)", lambda r: f"{max((e.duration_min for e in r['orbit'].eclipse_events), default=0):.1f}"),
    ("Contacts/Day",      lambda r: f"{r['orbit'].contacts_per_day:.1f}"),
    ("Contact Time (min/d)", lambda r: f"{r['orbit'].contact_time_per_day_min:.1f}"),
    ("Temp Max (°C)",     lambda r: f"{max(r['thermal'].node_temps_max.values(), default=0):.1f}"),
    ("Temp Min (°C)",     lambda r: f"{min(r['thermal'].node_temps_min.values(), default=0):.1f}"),
    ("Battery DOD (%)",   lambda r: f"{r['budget'].battery_dod_pct:.1f}"),
    ("Power Margin (W)",  lambda r: f"{r['budget'].power_margin_w:.1f}"),
    ("TID 5yr (krad)",    lambda r: f"{r['radiation'].tid_krad_5yr:.1f}"),
    ("Data/Day (GB)",     lambda r: f"{r['budget'].data_per_day_gb:.1f}"),
    ("Total Score",       lambda r: f"{r['score'].total_score:.0f}"),
    ("Grade",             lambda r: str(r['score'].grade)),
]


def _col_headers(results_history, scenario_names):
    return [(scenario_names[i] if i < len(scenario_names) else f"SAT-{i+1}")
            for i in range(len(results_history))]


def export_excel(results_history: list, scenario_names: list, path: str) -> None:
    """openpyxl 로 분석 결과 Excel 저장"""
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mission Analysis"

    # ── 색상 정의 ────────────────────────────────────────────────────────────
    DARK    = "0A0F1E"
    HEADER  = "0D2030"
    COL_HDR = "00DCFF"
    ROW_HDR = "1E3A4A"
    WHITE   = "C8E0F0"
    GRADE_G = "39FF96"
    GRADE_Y = "FFDC40"
    GRADE_R = "FF6B6B"

    thin = Side(style="thin", color="1E3A4A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell_style(cell, bg=DARK, fg=WHITE, bold=False, align="center"):
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.font      = Font(color=fg, bold=bold, name="Consolas", size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = border

    # ── 제목 행 ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{chr(66 + len(results_history))}1")
    title_cell = ws.cell(1, 1, f"SpaceD-AADE  Mission Analysis Export  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    cell_style(title_cell, bg=HEADER, fg=COL_HDR, bold=True)
    ws.row_dimensions[1].height = 22

    # ── 열 헤더 (시나리오 이름) ───────────────────────────────────────────────
    col_names = _col_headers(results_history, scenario_names)
    ws.cell(2, 1, "Metric")
    cell_style(ws.cell(2, 1), bg=ROW_HDR, fg=COL_HDR, bold=True)
    for ci, name in enumerate(col_names, start=2):
        c = ws.cell(2, ci, name)
        cell_style(c, bg=ROW_HDR, fg=COL_HDR, bold=True)
    ws.row_dimensions[2].height = 18

    # ── 데이터 ────────────────────────────────────────────────────────────────
    for ri, (label, func) in enumerate(EXPORT_ROWS, start=3):
        ws.cell(ri, 1, label)
        cell_style(ws.cell(ri, 1), bg=HEADER, fg=WHITE, align="left")
        for ci, results in enumerate(results_history, start=2):
            val = func(results)
            c = ws.cell(ri, ci, val)
            # Grade 열 색상
            if label == "Grade":
                fg = GRADE_G if "A" in val else (GRADE_R if "F" in val else GRADE_Y)
                cell_style(c, fg=fg, bold=True)
            else:
                cell_style(c)
        ws.row_dimensions[ri].height = 16

    # ── 열 너비 자동 조정 ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    for ci in range(2, len(results_history) + 2):
        ws.column_dimensions[chr(64 + ci)].width = 14

    wb.save(path)


def export_pdf(results_history: list, scenario_names: list, path: str) -> None:
    """Qt QPrinter + QTextDocument 로 HTML→PDF 저장"""
    from PySide6.QtPrintSupport import QPrinter
    from PySide6.QtGui          import QTextDocument
    from PySide6.QtCore         import QMarginsF

    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(path)
    printer.setPageMargins(QMarginsF(15, 15, 15, 15), printer.pageLayout().units())

    col_names = _col_headers(results_history, scenario_names)

    # ── HTML 빌드 ─────────────────────────────────────────────────────────────
    th_style = "background:#0d2030;color:#00dcff;padding:5px 8px;border:1px solid #1e3a4a;font-weight:bold;"
    td_style = "background:#050a14;color:#c8e0f0;padding:4px 8px;border:1px solid #1e3a4a;text-align:center;"
    td_label_style = td_style.replace("text-align:center", "text-align:left") + "color:#8ab0c0;"

    header_row = "".join(f"<th style='{th_style}'>{n}</th>" for n in ["Metric"] + col_names)
    body_rows  = ""
    for label, func in EXPORT_ROWS:
        vals = "".join(f"<td style='{td_style}'>{func(r)}</td>" for r in results_history)
        body_rows += f"<tr><td style='{td_label_style}'>{label}</td>{vals}</tr>"

    html = f"""
    <html><head>
    <meta charset='utf-8'>
    <style>
      body {{ background:#020508; color:#c8e0f0; font-family:Consolas,monospace; font-size:9pt; }}
      h2   {{ color:#00dcff; letter-spacing:2px; margin-bottom:6px; }}
      p    {{ color:#4a6a7a; font-size:8pt; margin-top:0; }}
      table {{ border-collapse:collapse; width:100%; }}
    </style>
    </head><body>
    <h2>SpaceD-AADE  Mission Analysis Report</h2>
    <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} &nbsp;|&nbsp;
       Scenarios: {len(results_history)}</p>
    <table>
      <thead><tr>{header_row}</tr></thead>
      <tbody>{body_rows}</tbody>
    </table>
    </body></html>
    """

    doc = QTextDocument()
    doc.setHtml(html)
    doc.print_(printer)
